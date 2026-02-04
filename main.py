"""
Streamlit app for Advanced Cellular Template Processing

Place this file as app.py in your repository and deploy to Streamlit Community Cloud.
"""

import os
import io
import re
import json
import time
import zipfile
import shutil
import base64
import tempfile
import requests
from pathlib import Path
from typing import Optional, List, Tuple

import streamlit as st
import openpyxl
from PIL import Image, ImageOps, ImageEnhance
# ---------------- Configuration ----------------
API_BASE = "https://integrate.api.nvidia.com/v1"
MODEL_SERVICE_DEFAULT = "meta/llama-3.2-90b-vision-instruct"
MODEL_GENERIC_DEFAULT = "meta/llama-3.2-90b-vision-instruct"

# ---------------- Schemas ----------------
SERVICE_SCHEMA = {
    "nr_arfcn": "number",
    "nr_band": "number",
    "nr_pci": "number",
    "nr_bw": "number",
    "nr5g_rsrp": "number",
    "nr5g_rsrq": "number",
    "nr5g_sinr": "number",
    "lte_band": "number",
    "lte_earfcn": "number",
    "lte_pci": "number",
    "lte_bw": "number",
    "lte_rsrp": "number",
    "lte_rsrq": "number",
    "lte_sinr": "number",
}

GENERIC_SCHEMAS = {
    "speed_test": {
        "image_type": "speed_test",
        "data": {
            "download_mbps": "number",
            "upload_mbps": "number",
            "ping_ms": "number",
            "jitter_ms": "number",
        },
    },
    "video_test": {
        "image_type": "video_test",
        "data": {
            "max_resolution": "string",
            "load_time_ms": "number",
            "buffering_percentage": "number",
        },
    },
    "voice_call": {
        "image_type": "voice_call",
        "data": {
            "phone_number": "string",
            "call_duration_seconds": "number",
            "call_status": "string",
            "time": "string",
        },
    },
}

# ---------------- Globals (reinitialized per run) ----------------
alpha_service = {}
beta_service = {}
gamma_service = {}

alpha_speedtest = {}
beta_speedtest = {}
gamma_speedtest = {}

alpha_video = {}
beta_video = {}
gamma_video = {}

voice_test = {}
extract_text = []
avearge = {}

# ---------------- Helpers ----------------
def _api_headers(token: str) -> dict:
    """Generate headers for NVIDIA API."""
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }


def log_append(log_placeholder, logs_list: list, msg: str):
    """Append a timestamped log line and refresh the placeholder text area."""
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    logs_list.append(line)
    # keep last 2000 lines
    display = "\n".join(logs_list[-2000:])
    try:
        log_placeholder.text_area("Logs", value=display, height=360)
    except Exception:
        # fallback to stdout if placeholder fails
        print(line)


def get_sector_from_col(col_index: int) -> str:
    if 0 <= col_index < 4:
        return "alpha"
    if 4 <= col_index < 8:
        return "beta"
    if 8 <= col_index < 12:
        return "gamma"
    if 12 <= col_index < 18:
        return "voicetest"
    return "unknown"

#---------------- UPDATED CLEANER (Safety Net) ----------------------

def clean_json_response(content: str) -> str:
    """
    Surgical cleaning: Finds the first valid JSON object by counting braces.
    This handles cases where the model outputs multiple JSONs or chatty text.
    """
    if not content:
        return content
    
    content = content.strip()
    
    # Locate the first opening brace
    start_idx = content.find('{')
    if start_idx == -1:
        return content
    
    # Count braces to find the matching closing brace
    brace_count = 0
    in_string = False
    escape = False
    
    for i, char in enumerate(content[start_idx:], start=start_idx):
        if char == '"' and not escape:
            in_string = not in_string
        
        if not in_string:
            if char == '{':
                brace_count += 1
            elif char == '}':
                brace_count -= 1
                
                # We found the end of the first valid object
                if brace_count == 0:
                    return content[start_idx : i+1]
        
        # Handle escape characters for string parsing logic
        if char == '\\' and not escape:
            escape = True
        else:
            escape = False

    # Fallback: If counting failed (malformed), try strictly the regex approach
    match = re.search(r'(\{.*\})', content, re.DOTALL)
    if match:
        return match.group(1)
        
    return content

# ---------------- Image extraction (only .xlsx now) ----------------
def extract_images_from_excel(xlsx_path: str, output_folder: str, log_placeholder, logs: list) -> List[str]:
    log_append(log_placeholder, logs, f"[LOG] Analyzing template file: {xlsx_path}")
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb.active
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] Could not open/read Excel file: {e}")
        return []

    images = getattr(sheet, "_images", [])
    if not images:
        log_append(log_placeholder, logs, "[WARN] No images found in workbook.")
        return []

    os.makedirs(output_folder, exist_ok=True)
    images_with_locations = []
    for image in images:
        try:
            row = image.anchor._from.row + 1
            col = image.anchor._from.col
        except Exception:
            row, col = 0, 0
        images_with_locations.append({"image": image, "row": row, "col": col})

    images_sorted = sorted(images_with_locations, key=lambda i: (i["row"], i["col"]))
    saved_paths = []
    counters = {"alpha": 0, "beta": 0, "gamma": 0, "voicetest": 0, "unknown": 0}
    log_append(log_placeholder, logs, f"[LOG] Found {len(images_sorted)} images. Extracting...")

    for itm in images_sorted:
        sector = get_sector_from_col(itm["col"])
        counters[sector] += 1
        filename = f"{sector}_image_{counters[sector]}.png"
        out_path = os.path.join(output_folder, filename)
        try:
            img_data = itm["image"]._data()
            pil = Image.open(io.BytesIO(img_data))
            pil.save(out_path, "PNG")
            saved_paths.append(out_path)
            try:
                loc = f"{openpyxl.utils.get_column_letter(itm['col']+1)}{itm['row']}"
            except Exception:
                loc = ""
            log_append(log_placeholder, logs, f"  - Saved {filename} {loc}")
        except Exception as e:
            log_append(log_placeholder, logs, f"[ERROR] Failed to save {filename}: {e}")

    return saved_paths


# ---------------- API helpers & analyzers ----------------
def _post_chat_completion(token: str, payload: dict, timeout: int = 60):
    headers = _api_headers(token)
    return requests.post(url=f"{API_BASE}/chat/completions", headers=headers, data=json.dumps(payload), timeout=timeout)


def process_service_images(token: str, image1_path: str, image2_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    sector = Path(image1_path).stem.split("_")[0]
    log_append(log_placeholder, logs, f"[LOG] Processing Service Images for '{sector}' (Samsung Pixel-Perfect)...")
    
    try:
        # 1. LOAD & ENHANCE (High Contrast for Accuracy)
        img1 = Image.open(image1_path)
        img2 = Image.open(image2_path)
        
        def enhance_img(img):
            img = ImageOps.grayscale(img)
            enhancer = ImageEnhance.Contrast(img)
            img = enhancer.enhance(2.5) # High contrast makes text black/white
            enhancer = ImageEnhance.Sharpness(img)
            img = enhancer.enhance(2.0)
            return img

        img1 = enhance_img(img1)
        img2 = enhance_img(img2)

        # 2. VERTICAL STITCH (Top/Bottom)
        if img1.width != img2.width:
            ratio = img1.width / img2.width
            new_height = int(img2.height * ratio)
            img2 = img2.resize((img1.width, new_height), Image.Resampling.LANCZOS)

        total_height = img1.height + img2.height + 50
        max_width = img1.width
        
        stitched = Image.new('L', (max_width, total_height), 255)
        stitched.paste(img1, (0, 0))
        
        # Black separator line
        from PIL import ImageDraw
        draw = ImageDraw.Draw(stitched)
        draw.rectangle([0, img1.height, max_width, img1.height + 50], fill=0)
        
        stitched.paste(img2, (0, img1.height + 50))
        
        buf = io.BytesIO()
        stitched.save(buf, format='PNG')
        b64_stitched = base64.b64encode(buf.getvalue()).decode("utf-8")
        
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] Stitching failed: {e}")
        return None

    # 3. PIXEL-PERFECT PROMPT (Maps exactly to your 1.png and 2.png)
    prompt = (
        "You are a regex-based OCR engine for Samsung Service Mode.\n"
        "Extract values strictly based on the text patterns below.\n\n"
        "**LTE BLOCK (Top of Image):**\n"
        "1. `lte_band`: Look for 'LTE RRC:CONN BAND:'. The number immediately after 'BAND:'.\n"
        "2. `lte_bw`: On the same line as Band, look for 'BW:'.\n"
        "3. `lte_earfcn`: Look for the line starting with 'Earfcn:'. Take the FIRST number before the comma. (Do NOT take the PCI value).\n"
        "4. `lte_pci`: On the 'Earfcn' line, look for 'PCI:'.\n"
        "5. `lte_rsrp`: Look for 'RSRP:'. The number immediately after it.\n"
        "6. `lte_rsrq`: Look for 'RSRQ:'.\n"
        "7. `lte_sinr`: Look for 'SNR:'. (Note: Label is SNR, not SINR).\n\n"
        "**5G NR BLOCK (Middle/Bottom of Image):**\n"
        "1. `nr5g_rsrp`: Look for the exact line 'NR5G_RSRP :' or 'NR_ANT MAX RSRP:'. Take the value.\n"
        "2. `nr5g_sinr`: Look for the exact line 'NR5G_SINR :'.\n"
        "3. `nr5g_rsrq`: Look for the exact line 'NR5G RSRQ :' or 'NR5G_RSRQ :'.\n"
        "4. `nr_arfcn`: Look for 'NR_ARFCN:'.\n"
        "5. `nr_pci`: Look for 'NR_PCI:'.\n"
        "6. `nr_band`: Look for 'NR_BAND:'. If it says 'n77', extract 77.\n"
        "7. `nr_bw`: Look for 'NR_BW:'. Found usually in the bottom half.\n\n"
        "**FORMATTING:**\n"
        "- Return ONLY valid JSON.\n"
        "- If a value is not found, return null.\n"
        f"SCHEMA:\n{json.dumps(SERVICE_SCHEMA, indent=2)}"
    )

    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_stitched}"}},
                ],
            }
        ],
    }

    try:
        resp = _post_chat_completion(token, payload, timeout=120)
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        content = clean_json_response(content)
        result = json.loads(content)
        log_append(log_placeholder, logs, f"[SUCCESS] Extracted service data for '{sector}'.")
        return result
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] API call failed: {e}")
        return None
    finally:
        time.sleep(2)

# ---------------- SPECIALIZED ANALYZERS (With Telecom Domain Logic) ----------------

def analyze_speed_test(token: str, image_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    """
    Strictly looks for Speed Test data.
    Includes PYTHON-SIDE SANITY CHECKS to fix "558000" type errors.
    """
    try:
        with open(image_path, "rb") as f:
            b = base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        return None

    # UPDATED PROMPT
    prompt = (
        "You are a Senior RF Engineer extracting SPEED TEST metrics.\n"
        "Ignore the 'context' of the image (e.g. ads, logos) and focus ONLY on the data.\n\n"
        "CRITICAL EXTRACTION STEPS:\n"
        "1. **Find Numbers**: Look for the largest numbers (Download/Upload).\n"
        "2. **Check Units**: \n"
        "   - The image is likely in **Mbps**.\n"
        "   - Do **NOT** convert to Kbps. Return the number exactly as seen.\n"
        "   - If you see '558.00', return 558. Do NOT ignore the decimal.\n"
        "3. **Corner Scan**: Look for 'Ping' or 'Latency' in corners (e.g., 20 ms).\n\n"
        "REQUIRED OUTPUT:\n"
        "Return valid JSON matching this schema exactly. Use null for missing values.\n"
        f"SCHEMA:\n{json.dumps(GENERIC_SCHEMAS['speed_test'], indent=2)}"
    )

    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}, {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b}"}}]}]
    }

    try:
        resp = _post_chat_completion(token, payload, timeout=50)
        resp.raise_for_status()
        content = clean_json_response(resp.json()["choices"][0]["message"]["content"])
        res = json.loads(content)
        
        # --- PYTHON SANITY CHECKS (The Fix for 558000) ---
        data = res.get("data", {})
        
        # 1. Download Sanity (Max realistic 5G is ~5000 Mbps)
        dl = data.get("download_mbps")
        if dl is not None:
            if dl > 10000: # If > 10,000, it's definitely an error (likely Kbps or decimal error)
                log_append(log_placeholder, logs, f"[AUTO-FIX] Download {dl} is impossible. Assuming Kbps -> Mbps.")
                data["download_mbps"] = dl / 1000
            elif dl in [2160, 1080, 720]: # Video Resolution Trap
                return None

        # 2. Upload Sanity
        ul = data.get("upload_mbps")
        if ul is not None and ul > 5000:
            data["upload_mbps"] = ul / 1000

        # 3. Ping Sanity
        ping = data.get("ping_ms")
        if ping is not None and ping < 1: 
             data["ping_ms"] = ping * 1000

        # Logic Check: Must have at least one valid metric
        if data.get("download_mbps") is None and data.get("upload_mbps") is None and data.get("ping_ms") is None:
            return None
            
        res["data"] = data # Save back modified data
        return res
    except Exception:
        return None


def analyze_video_test(token: str, image_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    """Strictly looks for Video Test data."""
    try:
        with open(image_path, "rb") as f:
            b = base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        return None

    # UPDATED PROMPT: Forces distinction between Marketing labels (4K) and Data (2160p)
    prompt = (
        "You are a Senior RF Engineer extracting VIDEO STREAMING metrics.\n\n"
        "CRITICAL EXTRACTION RULES:\n"
        "1. **Semantic Filter**: Ignore 'Speed Test' numbers (Mbps). Focus ONLY on Video stats.\n"
        "2. **Resolution Priority**:\n"
        "   - Look for specific NUMBERS: '2160p', '1080p', '1440p', '720p'.\n"
        "   - '4K' and 'HD' are marketing labels. PREFER the number (e.g. '2160p') if visible.\n"
        "   - Only return '4K' if no numeric resolution is found.\n"
        "3. **Load Time**: Find values in 'ms' (milliseconds). If 's' (1.2s), convert to ms (1200).\n\n"
        "REQUIRED OUTPUT:\n"
        "Return valid JSON matching this schema exactly.\n"
        f"SCHEMA:\n{json.dumps(GENERIC_SCHEMAS['video_test'], indent=2)}"
    )

    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}, {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b}"}}]}]
    }

    try:
        resp = _post_chat_completion(token, payload, timeout=50)
        resp.raise_for_status()
        content = clean_json_response(resp.json()["choices"][0]["message"]["content"])
        res = json.loads(content)
        
        if res.get("data", {}).get("max_resolution") is not None or res.get("data", {}).get("load_time_ms") is not None:
            return res
        return None
    except Exception:
        return None


def analyze_voice_test_strict(token: str, image_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    """Strictly looks for Voice Call data."""
    try:
        with open(image_path, "rb") as f:
            b = base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        return None

    # UPDATED PROMPT: Forces attention to the Timer
    prompt = (
        "You are a Telecom Engineer extracting VOICE CALL metrics.\n\n"
        "CRITICAL EXTRACTION RULES:\n"
        "1. **Visual Scan**: Look for a 'Dialer', 'Incoming Call', or 'Green Phone Icon'.\n"
        "2. **Timer Focus**: Locate the call timer (e.g. 00:12, 0:05).\n"
        "   - **Math**: Convert '00:12' -> 12 seconds.\n"
        "   - **Bias Check**: Do NOT assume duration is 0. Read the actual pixels. '00:04' is 4 seconds, not 0.\n"
        "3. **Ignore**: Speed/Video data.\n\n"
        "REQUIRED OUTPUT:\n"
        "Return valid JSON matching this schema exactly.\n"
        f"SCHEMA:\n{json.dumps(GENERIC_SCHEMAS['voice_call'], indent=2)}"
    )

    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}, {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b}"}}]}]
    }

    try:
        resp = _post_chat_completion(token, payload, timeout=50)
        resp.raise_for_status()
        content = clean_json_response(resp.json()["choices"][0]["message"]["content"])
        return json.loads(content)
    except Exception:
        return None

def dispatch_image_analysis(token: str, image_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    """
    Smart Router: Tries tests in a specific order based on filename index.
    """
    path_obj = Path(image_path)
    image_name = path_obj.stem
    
    # Extract index from filename (e.g. alpha_image_3 -> 3)
    try:
        idx = int(image_name.split("_")[-1])
    except:
        idx = 0

    log_append(log_placeholder, logs, f"[LOG] Dispatching '{image_name}'...")

    # HEURISTIC: Images 3-7 are usually Speed. Images 8+ are usually Video.
    if 3 <= idx <= 7:
        priority = ["speed", "video", "voice"]
    elif idx >= 8:
        priority = ["video", "speed", "voice"]
    else:
        priority = ["speed", "video", "voice"]

    for test_type in priority:
        res = None
        if test_type == "speed":
            res = analyze_speed_test(token, image_path, model_name, log_placeholder, logs)
        elif test_type == "video":
            res = analyze_video_test(token, image_path, model_name, log_placeholder, logs)
        elif test_type == "voice":
            res = analyze_voice_test_strict(token, image_path, model_name, log_placeholder, logs)
        
        if res:
            log_append(log_placeholder, logs, f"[SUCCESS] '{image_name}' identified as {test_type.upper()}.")
            return res

    log_append(log_placeholder, logs, f"[WARN] Could not identify '{image_name}' (tried {priority}).")
    return None

# Wrappers to maintain compatibility if called elsewhere
def evaluate_generic_image(token: str, image_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    return dispatch_image_analysis(token, image_path, model_name, log_placeholder, logs)

def evaluate_voice_image(token: str, image_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    return analyze_voice_test_strict(token, image_path, model_name, log_placeholder, logs)
# ---------------- Careful evaluation functions ----------------
def evaluate_service_images(token: str, image1_path: str, image2_path: str, model_name: str, log_placeholder, logs: list) -> Optional[dict]:
    sector = Path(image1_path).stem.split("_")[0] if image1_path else "unknown"
    log_append(log_placeholder, logs, f"[EVAL] Re-evaluating service images for '{sector}' (Vertical Smart Stitch)...")
    
    # 1. SMART VERTICAL STITCHING & ENHANCEMENT
    try:
        # Load images
        img1 = Image.open(image1_path)
        img2 = Image.open(image2_path)
        
        # Pre-process: Convert to Greyscale & Enhance Contrast (Make text POP)
        def enhance_img(img):
            img = ImageOps.grayscale(img)
            enhancer = ImageEnhance.Contrast(img)
            img = enhancer.enhance(2.0) # High contrast
            # Optional: Slight sharpness boost to make text edges crisp
            enhancer = ImageEnhance.Sharpness(img)
            img = enhancer.enhance(1.5)
            return img

        img1 = enhance_img(img1)
        img2 = enhance_img(img2)

        # Vertical Stack (Top/Bottom)
        # Resize width of img2 to match img1 to keep alignment
        if img1.width != img2.width:
            ratio = img1.width / img2.width
            new_height = int(img2.height * ratio)
            img2 = img2.resize((img1.width, new_height), Image.Resampling.LANCZOS)

        total_height = img1.height + img2.height + 50 # 50px padding
        max_width = img1.width
        
        # Create canvas (White background)
        stitched = Image.new('L', (max_width, total_height), 255)
        
        # Paste Image 1 (Top)
        stitched.paste(img1, (0, 0))
        
        # Draw a black separator line for clarity
        from PIL import ImageDraw
        draw = ImageDraw.Draw(stitched)
        draw.rectangle([0, img1.height, max_width, img1.height + 50], fill=0)
        
        # Paste Image 2 (Bottom)
        stitched.paste(img2, (0, img1.height + 50))
        
        # Save to buffer
        buf = io.BytesIO()
        stitched.save(buf, format='PNG')
        b64_stitched = base64.b64encode(buf.getvalue()).decode("utf-8")
        
    except Exception as e:
        log_append(log_placeholder, logs, f"[EVAL ERROR] Could not stitch/encode images: {e}")
        return None

    # 2. SAMSUNG OCR PROMPT (The "100% Accuracy" Fix)
    # This prompt maps the EXACT text labels from your image to the JSON keys.
    prompt = (
        "You are an Optical Character Recognition (OCR) engine for Samsung Service Mode screens.\n"
        "Analyze the stitched image. Extract values based on these EXACT text labels:\n\n"
        "**LTE SECTION (Look for 'Serving PLMN...LTE'):**\n"
        "- `lte_band`: Value after 'BAND:' (e.g., 66, 5)\n"
        "- `lte_bw`: Value after 'BW:' (e.g., 10, 15)\n"
        "- `lte_earfcn`: Value after 'Earfcn:'\n"
        "- `lte_pci`: Value after 'PCI:'\n"
        "- `lte_rsrp`: Value after 'RSRP:'\n"
        "- `lte_rsrq`: Value after 'RSRQ:'\n"
        "- `lte_sinr`: Value after 'SNR:'\n\n"
        "**5G NR SECTION (Look for 'NR5G_' or 'NR_'):**\n"
        "- `nr5g_rsrp`: Value after 'NR5G_RSRP :' or 'NR_ANT MAX RSRP:'\n"
        "- `nr5g_sinr`: Value after 'NR5G_SINR :'\n"
        "- `nr5g_rsrq`: Value after 'NR5G RSRQ :' or 'NR_ANT MIN RSRP' (if RSRQ not found)\n"
        "- `nr_arfcn`: Value after 'NR_ARFCN:'\n"
        "- `nr_pci`: Value after 'NR_PCI:'\n"
        "- `nr_band`: Value after 'NR_BAND:' (extract number, e.g., n77 -> 77)\n"
        "- `nr_bw`: Value after 'NR_BW:'\n\n"
        "**RULES:**\n"
        "1. Return ONLY the JSON object. No other text.\n"
        "2. If a value is missing on screen, use null.\n"
        f"SCHEMA:\n{json.dumps(SERVICE_SCHEMA, indent=2)}"
    )

    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_stitched}"}},
                ],
            }
        ],
    }

    try:
        resp = _post_chat_completion(token, payload, timeout=120)
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        content = clean_json_response(content) 
        return json.loads(content)
    except Exception as e:
        log_append(log_placeholder, logs, f"[EVAL ERROR] Service evaluation failed: {e}")
        return None
    finally:
        log_append(log_placeholder, logs, "[EVAL] Cooldown: waiting 2 seconds")
        time.sleep(2)

# ---------------- Expression resolution helpers ----------------
key_pattern = re.compile(r"\[['\"]([^'\"]+)['\"]\]")


def _normalize_name(s: str) -> str:
    return re.sub(r"[^0-9a-zA-Z]", "", s).lower()


def resolve_expression_with_vars(expr: str, allowed_vars: dict):
    expr = expr.strip()
    m = re.match(r"^([A-Za-z_]\w*)(.*)$", expr)
    if not m:
        return None
    base_raw = m.group(1)
    rest = m.group(2) or ""

    norm_map = {_normalize_name(k): k for k in allowed_vars.keys()}
    base_key = norm_map.get(_normalize_name(base_raw))
    if not base_key:
        for k in allowed_vars.keys():
            if k.lower() == base_raw.lower():
                base_key = k
                break
    if not base_key:
        return None

    obj = allowed_vars[base_key]
    if rest.strip() == "":
        return obj

    keys = key_pattern.findall(rest)
    if not keys:
        return None

    try:
        for k in keys:
            if not isinstance(obj, dict):
                return None
            if k in obj:
                obj = obj[k]
                continue
            found = None
            for real_k in obj.keys():
                if real_k.lower() == k.lower() or _normalize_name(real_k) == _normalize_name(k):
                    found = real_k
                    break
            if found:
                obj = obj[found]
            else:
                return None
        return obj
    except Exception:
        return None


def set_nested_value_case_insensitive(target: dict, keys: list, value):
    cur = target
    for idx, k in enumerate(keys):
        last = idx == (len(keys) - 1)
        if last:
            if isinstance(cur, dict):
                found = None
                for real_k in list(cur.keys()):
                    if real_k.lower() == k.lower() or _normalize_name(real_k) == _normalize_name(k):
                        found = real_k
                        break
                if found:
                    cur[found] = value
                else:
                    cur[k] = value
            return True
        else:
            found = None
            if isinstance(cur, dict):
                for real_k in list(cur.keys()):
                    if real_k.lower() == k.lower() or _normalize_name(real_k) == _normalize_name(k):
                        found = real_k
                        break
            if found:
                if not isinstance(cur[found], dict):
                    cur[found] = {}
                cur = cur[found]
            else:
                cur[k] = {}
                cur = cur[k]
    return True


def ask_model_for_expression_value(token: str, var_name: str, var_obj, expression: str, model_name: str, log_placeholder, logs: list):
    """Ask model to evaluate expression using only provided JSON variable; return value or None."""
    try:
        var_json = json.dumps(var_obj, indent=2)
    except Exception:
        var_json = json.dumps(str(var_obj))

    # UPDATED PROMPT: STRICT FORMATTING
    prompt = (
        f"You are an exact assistant. You are given a JSON variable named '{var_name}':\n\n"
        f"{var_json}\n\nGiven the expression:\n{expression}\n\n"
        "Using ONLY the provided JSON variable, return exactly one JSON object:\n{ \"value\": <value> }\n"
        "Where <value> is the exact value or null. "
        "STRICTLY return ONLY the JSON object. Do not add conversational text. "
        "Start your response with '{' and end with '}'."
    )

    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}]}],
    }

    try:
        resp = _post_chat_completion(token, payload, timeout=30)
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        content = clean_json_response(content)
        parsed = json.loads(content)
        return parsed.get("value", None)
    except Exception as e:
        log_append(log_placeholder, logs, f"[ASK-MODEL] Failed for expr {expression}: {e}")
        if "resp" in locals():
            log_append(log_placeholder, logs, f"  Response: {getattr(resp, 'text', '')}")
        return None
# ---------------- Main processing function for Streamlit ----------------
def process_file_streamlit(user_file_path: str,
                           token: str,
                           temp_dir: str,
                           logs: list,
                           text_area_placeholder,
                           model_service: str = MODEL_SERVICE_DEFAULT,
                           model_generic: str = MODEL_GENERIC_DEFAULT) -> Optional[str]:
    """
    Main worker. IMPORTANT: user_file_path is expected to be a local filesystem path.
    """

    # reinitialize globals
    global alpha_service, beta_service, gamma_service
    global alpha_speedtest, beta_speedtest, gamma_speedtest
    global alpha_video, beta_video, gamma_video
    global voice_test, extract_text, avearge

    alpha_service = {}
    beta_service = {}
    gamma_service = {}

    alpha_speedtest = {}
    beta_speedtest = {}
    gamma_speedtest = {}

    alpha_video = {}
    beta_video = {}
    gamma_video = {}

    voice_test = {}
    extract_text = []
    avearge = {}

    # ensure temp dir exists
    os.makedirs(temp_dir, exist_ok=True)
    images_temp = os.path.join(temp_dir, "images")
    os.makedirs(images_temp, exist_ok=True)

    # Use provided path directly
    local_template = user_file_path
    if not os.path.exists(local_template):
        log_append(text_area_placeholder, logs, f"[ERROR] Template not found: {local_template}")
        return None

    # only support .xlsx
    path_obj = Path(local_template)
    if path_obj.suffix.lower() != ".xlsx":
        log_append(text_area_placeholder, logs, "[ERROR] Unsupported file type (only .xlsx supported now).")
        return None

    # extract images from workbook
    image_paths = extract_images_from_excel(local_template, images_temp, text_area_placeholder, logs)
    if not image_paths:
        log_append(text_area_placeholder, logs, "[ERROR] No images to process (workbook may not contain images).")
        return None

    # group images by sector
    images_by_sector = {"alpha": [], "beta": [], "gamma": [], "voicetest": [], "unknown": []}
    for p in image_paths:
        sector = Path(p).stem.split("_")[0]
        if sector in images_by_sector:
            images_by_sector[sector].append(p)
        else:
            images_by_sector["unknown"].append(p)

    log_append(text_area_placeholder, logs, "[LOG] Starting main processing loop.")
    
    # --- MAIN LOOP (Alpha, Beta, Gamma) ---
    for sector in ["alpha", "beta", "gamma"]:
        log_append(text_area_placeholder, logs, f"--- Processing sector: {sector.upper()} ---")
        sector_images = images_by_sector[sector]

        # 1. Process Service Images
        img1 = next((p for p in sector_images if Path(p).stem.endswith("_image_1")), None)
        img2 = next((p for p in sector_images if Path(p).stem.endswith("_image_2")), None)

        if img1 and img2:
            svc = process_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
            if svc:
                if sector == "alpha":
                    alpha_service = svc
                elif sector == "beta":
                    beta_service = svc
                elif sector == "gamma":
                    gamma_service = svc
        else:
            log_append(text_area_placeholder, logs, f"[WARN] Missing service images for {sector}")

        # 2. Process Other Images (Speed/Video/Voice) using Dispatcher
        other_images = [
            p for p in sector_images
            if not (Path(p).stem.endswith("_image_1") or Path(p).stem.endswith("_image_2"))
        ]
        
        for img in other_images:
            # CHANGED: Use Smart Dispatcher
            res = dispatch_image_analysis(token, img, model_generic, text_area_placeholder, logs)
            
            if res and "image_type" in res:
                image_name = Path(img).stem
                if res["image_type"] == "speed_test":
                    if sector == "alpha":
                        alpha_speedtest[image_name] = res.get("data", {})
                    elif sector == "beta":
                        beta_speedtest[image_name] = res.get("data", {})
                    elif sector == "gamma":
                        gamma_speedtest[image_name] = res.get("data", {})
                elif res["image_type"] == "video_test":
                    if sector == "alpha":
                        alpha_video[image_name] = res.get("data", {})
                    elif sector == "beta":
                        beta_video[image_name] = res.get("data", {})
                    elif sector == "gamma":
                        gamma_video[image_name] = res.get("data", {})
                elif res["image_type"] == "voice_call":
                    voice_test[image_name] = res.get("data", {})

    # --- MAIN LOOP (Voicetest Sector) ---
    if images_by_sector["voicetest"]:
        log_append(text_area_placeholder, logs, "--- Processing sector: VOICETEST ---")
        for img in images_by_sector["voicetest"]:
            # CHANGED: Use Strict Voice Analyzer
            res = analyze_voice_test_strict(token, img, model_generic, text_area_placeholder, logs)
            if res and "data" in res:
                voice_test[Path(img).stem] = res.get("data", {})

    # ---------------- Evaluation pass & Rule 2 ----------------
    log_append(text_area_placeholder, logs, "\n[LOG] Starting evaluation pass to refill missing/null fields.")
    retried_service_sectors = set()
    retried_images = set()

    def contains_nulls(d):
        if not isinstance(d, dict):
            return False
        for v in d.values():
            if v is None:
                return True
            if isinstance(v, dict) and contains_nulls(v):
                return True
        return False

    # Evaluate service dicts
    for sector in ["alpha", "beta", "gamma"]:
        svc_var = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
        if not svc_var:
            img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
            img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
            if img1 and img2 and sector not in retried_service_sectors:
                log_append(text_area_placeholder, logs, f"[EVAL] Service dict empty for {sector}. Re-evaluating.")
                eval_res = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                retried_service_sectors.add(sector)
                if eval_res:
                    if sector == "alpha":
                        alpha_service = eval_res
                    elif sector == "beta":
                        beta_service = eval_res
                    elif sector == "gamma":
                        gamma_service = eval_res
            continue

        if contains_nulls(svc_var) and sector not in retried_service_sectors:
            img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
            img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
            if img1 and img2:
                log_append(text_area_placeholder, logs, f"[EVAL] Found nulls in {sector}_service; re-evaluating.")
                eval_res = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                retried_service_sectors.add(sector)
                if eval_res:
                    target = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
                    for k, v in eval_res.items():
                        if target.get(k) is None and v is not None:
                            target[k] = v

    # UPDATED HELPER: Retry single images using Smart Dispatcher
    def _retry_image_and_merge(image_name: str, sector_var_map: dict) -> bool:
        image_path = os.path.join(images_temp, f"{image_name}.png")
        
        # 1. Find file if not at exact path
        if not os.path.exists(image_path):
            found = None
            for s_list in images_by_sector.values():
                for p in s_list:
                    if Path(p).stem == image_name:
                        found = p
                        break
                if found:
                    break
            if found:
                image_path = found
            else:
                log_append(text_area_placeholder, logs, f"[EVAL WARN] Image {image_name} not found. Skipping.")
                return False
        
        # 2. Prevent double retry
        if image_path in retried_images:
            return False

        is_voice = image_name.startswith("voicetest")
        log_append(text_area_placeholder, logs, f"[EVAL] Retrying analysis for {image_name}.")
        
        res = None
        if is_voice:
            res = analyze_voice_test_strict(token, image_path, model_generic, text_area_placeholder, logs)
        else:
            # Use smart dispatcher for generic (Alpha/Beta/Gamma) images
            res = dispatch_image_analysis(token, image_path, model_generic, text_area_placeholder, logs)

        retried_images.add(image_path)
        
        if res and "data" in res:
            sector_var_map.setdefault(image_name, {})
            data = res.get("data", {})
            for k, v in data.items():
                if sector_var_map[image_name].get(k) is None and v is not None:
                    sector_var_map[image_name][k] = v
            return True
        return False

    sector_maps = [
        ("alpha", alpha_speedtest, alpha_video),
        ("beta", beta_speedtest, beta_video),
        ("gamma", gamma_speedtest, gamma_video),
    ]

    expected_indices = {"service": [1, 2], "speed": [3, 4, 5, 6, 7], "video": [8]}

    def missing_service_fields(svc_obj):
        missing = []
        for k in SERVICE_SCHEMA.keys():
            if k not in svc_obj or svc_obj.get(k) is None:
                missing.append(k)
        return missing

    for sector, speed_map, video_map in sector_maps:
        log_append(text_area_placeholder, logs, f"[RULE2] Verifying expected items for {sector}.")
        svc_var = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
        svc_missing = missing_service_fields(svc_var) if svc_var else list(SERVICE_SCHEMA.keys())
        if svc_missing:
            log_append(text_area_placeholder, logs, f"[RULE2] Service for {sector} missing: {svc_missing}")
            img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
            img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
            if img1 and img2 and sector not in retried_service_sectors:
                log_append(text_area_placeholder, logs, f"[RULE2] Re-process service images for {sector}.")
                normal_svc = process_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                retried_service_sectors.add(sector)
                if normal_svc:
                    target = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
                    for k, v in normal_svc.items():
                        if target.get(k) is None and v is not None:
                            target[k] = v
                    if missing_service_fields(target):
                        log_append(text_area_placeholder, logs, f"[RULE2] Careful eval for service images {sector}.")
                        eval_svc = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                        if eval_svc:
                            for k, v in eval_svc.items():
                                if target.get(k) is None and v is not None:
                                    target[k] = v
            else:
                log_append(text_area_placeholder, logs, f"[RULE2] Cannot re-process {sector} (missing or already retried).")

        for idx in expected_indices["speed"]:
            name = f"{sector}_image_{idx}"
            if name not in speed_map:
                log_append(text_area_placeholder, logs, f"[RULE2] Missing expected speed image {name}. Searching files.")
                file_path = next((p for p in images_by_sector[sector] if Path(p).stem == name), None)
                if file_path:
                    log_append(text_area_placeholder, logs, f"[RULE2] Found {name}. Processing.")
                    _retry_image_and_merge(name, speed_map)
                else:
                    log_append(text_area_placeholder, logs, f"[RULE2] No file for expected {name}.")
            else:
                missing = []
                for k in GENERIC_SCHEMAS["speed_test"]["data"].keys():
                    if k not in speed_map[name] or speed_map[name].get(k) is None:
                        missing.append(k)
                if missing:
                    log_append(text_area_placeholder, logs, f"[RULE2] {name} missing {missing}. Re-evaluating.")
                    _retry_image_and_merge(name, speed_map)

        for idx in expected_indices["video"]:
            name = f"{sector}_image_{idx}"
            if name not in video_map:
                file_path = next((p for p in images_by_sector[sector] if Path(p).stem == name), None)
                if file_path:
                    log_append(text_area_placeholder, logs, f"[RULE2] Found video {name}. Processing.")
                    _retry_image_and_merge(name, video_map)
                else:
                    log_append(text_area_placeholder, logs, f"[RULE2] No file for expected video {name}.")
            else:
                missing = []
                for k in GENERIC_SCHEMAS["video_test"]["data"].keys():
                    if k not in video_map[name] or video_map[name].get(k) is None:
                        missing.append(k)
                if missing:
                    log_append(text_area_placeholder, logs, f"[RULE2] {name} missing {missing}. Re-evaluating.")
                    _retry_image_and_merge(name, video_map)

    # voicetest checks
    log_append(text_area_placeholder, logs, "[RULE2] Verifying voicetest completeness.")
    for idx in [1, 2, 3]:
        name = f"voicetest_image_{idx}"
        if name not in voice_test:
            file_path = next((p for p in images_by_sector["voicetest"] if Path(p).stem == name), None)
            if file_path:
                log_append(text_area_placeholder, logs, f"[RULE2] Missing voice entry {name}. Processing.")
                _retry_image_and_merge(name, voice_test)
            else:
                log_append(text_area_placeholder, logs, f"[RULE2] No file for expected voice {name}.")
        else:
            missing = []
            for k in GENERIC_SCHEMAS["voice_call"]["data"].keys():
                if k not in voice_test[name] or voice_test[name].get(k) is None:
                    missing.append(k)
            if missing:
                log_append(text_area_placeholder, logs, f"[RULE2] {name} missing {missing}. Re-evaluating.")
                _retry_image_and_merge(name, voice_test)

    log_append(text_area_placeholder, logs, "[LOG] Rule 2 verification complete.")

    # ---------- compute averages ----------
    def _to_number(v):
        try:
            if v is None:
                return None
            if isinstance(v, bool):
                return None
            return float(v)
        except Exception:
            return None

    def _compute_speed_averages(speed_map):
        metrics = {"download_mbps": [], "upload_mbps": [], "ping_ms": []}
        for entry in speed_map.values():
            if not isinstance(entry, dict):
                continue
            for m in metrics.keys():
                val = _to_number(entry.get(m))
                if val is not None:
                    metrics[m].append(val)
        result = {}
        for m, vals in metrics.items():
            if vals:
                result[m] = sum(vals) / len(vals)
            else:
                result[m] = None
        return result

    avearge = {
        "avearge_alpha_speedtest": _compute_speed_averages(alpha_speedtest),
        "avearge_beta_speedtest": _compute_speed_averages(beta_speedtest),
        "avearge_gamma_speedtest": _compute_speed_averages(gamma_speedtest),
    }

    # ---------------- Mapping: extract bold+red expressions and replace ----------------
    log_append(text_area_placeholder, logs, "[LOG] Scanning workbook for BOLD+RED expressions and replacing with values.")
    try:
        wb_edit = openpyxl.load_workbook(local_template)
        sheet_edit = wb_edit.active
        cells_to_process = []

        def _font_is_strict_red(font):
            if not font:
                return False
            if not getattr(font, "bold", False):
                return False
            col = getattr(font, "color", None)
            if col is None:
                return False
            rgb = getattr(col, "rgb", None)
            if not rgb:
                return False
            up = str(rgb).upper()
            return up[-6:] == "FF0000"

        def _normalize_expr(raw: str) -> str:
            s = raw.strip()
            if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
                s = s[1:-1].strip()
            return s

        for row in sheet_edit.iter_rows(min_row=1, max_row=sheet_edit.max_row, min_col=1, max_col=16):
            for cell in row:
                val = cell.value
                if not val or not isinstance(val, str):
                    continue
                font = cell.font
                if not font:
                    continue
                if _font_is_strict_red(font):
                    expr = _normalize_expr(val)
                    if expr:
                        extract_text.append(expr)
                        cells_to_process.append((cell, expr))

        allowed_vars = {
            "alpha_service": alpha_service,
            "beta_service": beta_service,
            "gamma_service": gamma_service,
            "alpha_speedtest": alpha_speedtest,
            "beta_speedtest": beta_speedtest,
            "gamma_speedtest": gamma_speedtest,
            "alpha_video": alpha_video,
            "beta_video": beta_video,
            "gamma_video": gamma_video,
            "voice_test": voice_test,
            "avearge": avearge,
        }

        def _to_number_convert(v):
            try:
                if v is None:
                    return None
                if isinstance(v, (int, float)):
                    return v
                if isinstance(v, bool):
                    return None
                s = str(v).strip()
                s_clean = s.replace(",", "")
                if s_clean == "":
                    return None
                if re.fullmatch(r"[-+]?\d+", s_clean):
                    return int(s_clean)
                if re.fullmatch(r"[-+]?\d*\.\d+", s_clean):
                    return float(s_clean)
                return None
            except Exception:
                return None

        for cell_obj, expr in cells_to_process:
            resolved = resolve_expression_with_vars(expr, allowed_vars)
            if resolved is None:
                cell_obj.value = "NULL"
            else:
                if isinstance(resolved, str):
                    conv = _to_number_convert(resolved)
                    if conv is not None:
                        cell_obj.value = conv
                    else:
                        cell_obj.value = resolved
                elif isinstance(resolved, (int, float)):
                    cell_obj.value = resolved
                elif isinstance(resolved, (dict, list)):
                    try:
                        cell_obj.value = json.dumps(resolved)
                    except Exception:
                        cell_obj.value = str(resolved)
                else:
                    cell_obj.value = str(resolved)

        wb_edit.save(local_template)
        log_append(text_area_placeholder, logs, f"[LOG] Workbook updated and saved: {local_template}")
    except Exception as e:
        log_append(text_area_placeholder, logs, f"[ERROR] Failed mapping pass: {e}")

    # ---------------- Rule 3: remap NULL cells using strict AI re-checks ----------------
    log_append(text_area_placeholder, logs, "[LOG] Running Rule 3: remap any remaining NULL bold+red expressions with AI.")
    try:
        wb_r3 = openpyxl.load_workbook(local_template)
        sheet_r3 = wb_r3.active

        allowed_vars = {
            "alpha_service": alpha_service,
            "beta_service": beta_service,
            "gamma_service": gamma_service,
            "alpha_speedtest": alpha_speedtest,
            "beta_speedtest": beta_speedtest,
            "gamma_speedtest": gamma_speedtest,
            "alpha_video": alpha_video,
            "beta_video": beta_video,
            "gamma_video": gamma_video,
            "voice_test": voice_test,
            "avearge": avearge,
        }

        problematic_cells = []
        for row in sheet_r3.iter_rows(min_row=1, max_row=sheet_r3.max_row, min_col=1, max_col=16):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                if val.strip().upper() != "NULL":
                    continue
                font = cell.font
                if font and _font_is_strict_red(font):
                    problematic_cells.append(cell)

        remapped = 0
        for cell in problematic_cells:
            # find a candidate expression from extract_text that references a known base var
            candidate = None
            for ex in extract_text:
                mm = re.match(r"^([A-Za-z_]\w*)", ex.strip())
                if not mm:
                    continue
                base = mm.group(1)
                # normalize base
                if _normalize_name(base) in {_normalize_name(k) for k in allowed_vars.keys()}:
                    if resolve_expression_with_vars(ex, allowed_vars) is None:
                        candidate = ex
                        break
            if not candidate:
                continue

            expr = candidate
            log_append(text_area_placeholder, logs, f"[RULE3] Attempting strict re-map '{expr}' for cell {cell.coordinate}")
            m = re.match(r"^([A-Za-z_]\w*)(.*)$", expr)
            if not m:
                log_append(text_area_placeholder, logs, f"[RULE3] Could not parse '{expr}'. Skipping.")
                continue
            base_raw = m.group(1)
            rest = m.group(2) or ""
            norm_map = {_normalize_name(k): k for k in allowed_vars.keys()}
            base_key = norm_map.get(_normalize_name(base_raw))
            if not base_key:
                log_append(text_area_placeholder, logs, f"[RULE3] Base '{base_raw}' not found. Skipping.")
                continue

            # If service variable, try re-evaluate images first
            if base_key in ("alpha_service", "beta_service", "gamma_service"):
                sector = base_key.split("_")[0]
                img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
                img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
                if img1 and img2:
                    log_append(text_area_placeholder, logs, f"[RULE3] Re-evaluating service images for {sector} (strict).")
                    svc_eval = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                    if svc_eval:
                        allowed_vars[base_key].update(svc_eval)
                        resolved_after = resolve_expression_with_vars(expr, allowed_vars)
                        if resolved_after is not None:
                            keys = key_pattern.findall(rest)
                            set_nested_value_case_insensitive(allowed_vars[base_key], keys, resolved_after)
                            if isinstance(resolved_after, (int, float)):
                                cell.value = resolved_after
                            elif isinstance(resolved_after, str):
                                cell.value = resolved_after
                            else:
                                try:
                                    cell.value = json.dumps(resolved_after)
                                except Exception:
                                    cell.value = str(resolved_after)
                            remapped += 1
                            continue
                log_append(text_area_placeholder, logs, f"[RULE3] Asking model for '{expr}' using '{base_key}'")
                value = ask_model_for_expression_value(token, base_key, allowed_vars[base_key], expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    keys = key_pattern.findall(rest)
                    set_nested_value_case_insensitive(allowed_vars[base_key], keys, value)
                    if isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str):
                        cell.value = value
                    else:
                        try:
                            cell.value = json.dumps(value)
                        except Exception:
                            cell.value = str(value)
                    remapped += 1
                    continue
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Model couldn't provide value for '{expr}'.")
                    continue

            # For non-service variables (speed/video/voice), expect image key as first bracket
            keys = key_pattern.findall(rest)
            if not keys:
                log_append(text_area_placeholder, logs, f"[RULE3] No keys in '{expr}'. Skipping.")
                continue
            image_key = keys[0]
            file_path = None
            for lst in images_by_sector.values():
                for p in lst:
                    if Path(p).stem == image_key:
                        file_path = p
                        break
                if file_path:
                    break

            if not file_path:
                log_append(text_area_placeholder, logs, f"[RULE3] No file for image '{image_key}'. Asking model with '{base_key}'.")
                value = ask_model_for_expression_value(token, base_key, allowed_vars[base_key], expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    set_nested_value_case_insensitive(allowed_vars[base_key], keys[1:], value)
                    if isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str):
                        cell.value = value
                    else:
                        try:
                            cell.value = json.dumps(value)
                        except Exception:
                            cell.value = str(value)
                    remapped += 1
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Could not remap '{expr}'.")
                continue

            # If file found: strict evaluate image
            if image_key.startswith("voicetest"):
                log_append(text_area_placeholder, logs, f"[RULE3] Strictly evaluating voice image '{image_key}'.")
                voice_eval = evaluate_voice_image(token, file_path, model_generic, text_area_placeholder, logs)
                if voice_eval and "data" in voice_eval:
                    voice_test.setdefault(image_key, {}).update(voice_eval["data"])
                    nested_keys = keys[1:]
                    resolved_after = resolve_expression_with_vars(expr, {**allowed_vars, "voice_test": voice_test})
                    if resolved_after is not None:
                        set_nested_value_case_insensitive(voice_test, nested_keys, resolved_after)
                        if isinstance(resolved_after, (int, float)):
                            cell.value = resolved_after
                        elif isinstance(resolved_after, str):
                            cell.value = resolved_after
                        else:
                            try:
                                cell.value = json.dumps(resolved_after)
                            except Exception:
                                cell.value = str(resolved_after)
                        remapped += 1
                        continue
                log_append(text_area_placeholder, logs, f"[RULE3] Asking model for '{expr}' using 'voice_test'.")
                value = ask_model_for_expression_value(token, "voice_test", voice_test, expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    set_nested_value_case_insensitive(voice_test, keys[1:], value)
                    cell.value = value if not isinstance(value, dict) else json.dumps(value)
                    remapped += 1
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Could not remap '{expr}' from voice image.")
                continue
            else:
                # generic image (speed/video) - USE DISPATCHER FOR STRICT EVAL
                log_append(text_area_placeholder, logs, f"[RULE3] Strictly evaluating generic image '{image_key}'.")
                # CHANGED: Use the Smart Dispatcher for Rule 3 retry as well
                gen_eval = dispatch_image_analysis(token, file_path, model_generic, text_area_placeholder, logs)
                
                if gen_eval and "data" in gen_eval:
                    pref = image_key.split("_")[0]
                    if pref == "alpha":
                        if gen_eval.get("image_type") == "speed_test":
                            alpha_speedtest.setdefault(image_key, {}).update(gen_eval["data"])
                        elif gen_eval.get("image_type") == "video_test":
                            alpha_video.setdefault(image_key, {}).update(gen_eval["data"])
                    elif pref == "beta":
                        if gen_eval.get("image_type") == "speed_test":
                            beta_speedtest.setdefault(image_key, {}).update(gen_eval["data"])
                        elif gen_eval.get("image_type") == "video_test":
                            beta_video.setdefault(image_key, {}).update(gen_eval["data"])
                    elif pref == "gamma":
                        if gen_eval.get("image_type") == "speed_test":
                            gamma_speedtest.setdefault(image_key, {}).update(gen_eval["data"])
                        elif gen_eval.get("image_type") == "video_test":
                            gamma_video.setdefault(image_key, {}).update(gen_eval["data"])

                    # attempt to resolve now
                    new_allowed = {
                        "alpha_service": alpha_service, "beta_service": beta_service, "gamma_service": gamma_service,
                        "alpha_speedtest": alpha_speedtest, "beta_speedtest": beta_speedtest, "gamma_speedtest": gamma_speedtest,
                        "alpha_video": alpha_video, "beta_video": beta_video, "gamma_video": gamma_video,
                        "voice_test": voice_test, "avearge": avearge,
                    }
                    resolved_after = resolve_expression_with_vars(expr, new_allowed)
                    if resolved_after is not None:
                        nested_keys = key_pattern.findall(rest)
                        set_nested_value_case_insensitive(new_allowed[base_key], nested_keys, resolved_after)
                        if isinstance(resolved_after, (int, float)):
                            cell.value = resolved_after
                        elif isinstance(resolved_after, str):
                            cell.value = resolved_after
                        else:
                            try:
                                cell.value = json.dumps(resolved_after)
                            except Exception:
                                cell.value = str(resolved_after)
                        remapped += 1
                        continue

                log_append(text_area_placeholder, logs, f"[RULE3] Asking model for '{expr}' using '{base_key}'.")
                value = ask_model_for_expression_value(token, base_key, allowed_vars[base_key], expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    nested_keys = key_pattern.findall(rest)
                    set_nested_value_case_insensitive(allowed_vars[base_key], nested_keys, value)
                    if isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str):
                        cell.value = value
                    else:
                        try:
                            cell.value = json.dumps(value)
                        except Exception:
                            cell.value = str(value)
                    remapped += 1
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Could not remap '{expr}'. Left as NULL.")

        wb_r3.save(local_template)
        log_append(text_area_placeholder, logs, f"[RULE3] Remapping complete. Cells remapped: {remapped}. Workbook saved.")
    except Exception as e:
        log_append(text_area_placeholder, logs, f"[ERROR] Rule 3 remapping failed: {e}")

    # Final logs of variables
    log_append(text_area_placeholder, logs, "=" * 60)
    log_append(text_area_placeholder, logs, "FINAL STRUCTURED DATA (post-eval/rule2/rule3):")

    def _pp(name, obj):
        try:
            s = json.dumps(obj, indent=2)
        except Exception:
            s = str(obj)
        log_append(text_area_placeholder, logs, f"\n{name}:\n{s}")

    _pp("alpha_service", alpha_service)
    _pp("beta_service", beta_service)
    _pp("gamma_service", gamma_service)
    _pp("alpha_speedtest", alpha_speedtest)
    _pp("beta_speedtest", beta_speedtest)
    _pp("gamma_speedtest", gamma_speedtest)
    _pp("alpha_video", alpha_video)
    _pp("beta_video", beta_video)
    _pp("gamma_video", gamma_video)
    _pp("voice_test", voice_test)
    _pp("avearge", avearge)
    _pp("extract_text", extract_text)

    # Return the updated workbook path
    return local_template


# ---------------- Streamlit UI ----------------
def validate_api_key(token: str) -> Tuple[bool, str]:
    # lightweight format check
    if not token or "nvapi" not in token: # Simple check for Nvidia key format, but optional
        # NOTE: Nvidia keys typically start with nvapi- but not strictly required by logic, just a check
        # We'll allow it if it's not empty for flexibility.
        if len(token) < 10:
             return False, "Token looks too short."
    return True, "Token looks valid (format check)."


def main_ui():
    st.set_page_config(page_title="Advanced Cellular Template Processor", layout="wide")
    st.title("Advanced Cellular Template Processor")
    st.write("Provide an NVIDIA API key in the sidebar and validate it. After validation you can upload an .xlsx template.")

    # sidebar: token & simple validation
    st.sidebar.header("API Key & Settings")
    token_input = st.sidebar.text_input("NVIDIA API token", type="password", placeholder="nvapi-...")
    if "logs" not in st.session_state:
        st.session_state["logs"] = []
    if "API_VALID" not in st.session_state:
        st.session_state["API_VALID"] = False
    if "NVIDIA_TOKEN" not in st.session_state:
        st.session_state["NVIDIA_TOKEN"] = ""

    if st.sidebar.button("Validate API key"):
        ok, msg = validate_api_key(token_input)
        if ok:
            st.session_state["API_VALID"] = True
            st.session_state["NVIDIA_TOKEN"] = token_input
            st.sidebar.success("API token stored in session (format validated).")
            st.session_state["logs"].append("[UI] API token stored (format validated).")
        else:
            st.session_state["API_VALID"] = False
            st.sidebar.error(f"Validation failed: {msg}")
            st.session_state["logs"].append("[UI] API token validation failed.")

    # logs area (neat box)
    log_placeholder = st.empty()
    current_logs = "\n".join(st.session_state["logs"][-2000:])
    log_placeholder.text_area("Logs", value=current_logs, height=360)

    # only allow uploading when validated
    if st.session_state.get("API_VALID", False):
        st.header("Upload Template (.xlsx only)")
        uploaded_file = st.file_uploader("Upload .xlsx template", type=["xlsx"], accept_multiple_files=False)

        model_service = st.selectbox("Model for SERVICE images", options=[MODEL_SERVICE_DEFAULT], index=0)
        model_generic = st.selectbox("Model for GENERIC images", options=[MODEL_GENERIC_DEFAULT], index=0)

        if uploaded_file:
            # use a unique temporary directory to avoid collisions
            tmp_dir = tempfile.mkdtemp(prefix="streamlit_")
            saved_template_path = os.path.join(tmp_dir, uploaded_file.name)
            with open(saved_template_path, "wb") as f:
                f.write(uploaded_file.read())
            st.success(f"Saved uploaded file: {uploaded_file.name}")
            st.info("Temporary directory created for this upload (isolated).")

            if st.button("Process file now"):
                st.session_state["logs"].append("[UI] Starting processing...")
                log_append(log_placeholder, st.session_state["logs"], "[UI] Starting processing...")
                out_path = process_file_streamlit(
                    user_file_path=saved_template_path,
                    token=st.session_state["NVIDIA_TOKEN"],
                    temp_dir=tmp_dir,
                    logs=st.session_state["logs"],
                    text_area_placeholder=log_placeholder,
                    model_service=model_service,
                    model_generic=model_generic,
                )

                if out_path:
                    st.success("Processing finished.")
                    with open(out_path, "rb") as f:
                        st.download_button("Download processed file", data=f, file_name=os.path.basename(out_path))
                else:
                    st.error("Processing failed. Check logs for details.")
    else:
        st.info("Please validate your API key in the sidebar before uploading files.")


if __name__ == "__main__":
    main_ui()
